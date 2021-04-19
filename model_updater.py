import os
import time
import datetime
from win32com.client import Dispatch
from openpyxl import load_workbook

def get_workbooks(dir_list):
	#folder = subprocess.Popen(r'explorer /select,"C:\"')
	#get local directory
	ext = ['xlsx', 'xlsm', 'xls', 'csv', 'xml']
	wb_dirs = []
	for dir_root in dir_list:
		try:
			[wb_dirs.append(os.path.join(dir_root,x)) for x in os.listdir(dir_root) if x.endswith(tuple(ext)) and x[0] != "~"]
		except FileNotFoundError as e:
			create_log(str(e))

	return wb_dirs

def create_log(message, console = True):
	file_name = 'log.txt'

	with open(file_name, 'a+') as f:
		if console is True:
			print(message)
		f.write(message+'\n')

def exit_message(message, exit_time):
	create_log(message+'\n') #seperates run instance logs
	time.sleep(exit_time)
	quit()
		
def get_directoires():

	file_name = 'directories.txt'
	dir_list = ['#INCLUDE ALL FOLDER LOCATIONS THAT CONTAIN THE EXCEL FILES WITH LINKED VARIABLES\n',
				'C:\\Users\\Folder Location X\n',
				'C:\\Users\\Folder Location Y']

	try:
		with open(file_name, 'r') as f:
			lines = f.readlines()
			dir_list = []
			for line in lines:
				if line[0] == '#': #allows for comments to be included
					continue
				if line[1:3] != ':\\':
					exit_message('Invalid links contained within directories. Please fix before running the application again.\n', 6)
					
				dir_list.append(line.strip('\n'))

	except FileNotFoundError:
		with open(file_name, 'x') as f:
			f.writelines(dir_list)

		exit_message('Created directories.txt where the application is being ran. Update file with the correct folder locations before running the application again.', 6)
	
	return dir_list

def modifyFormula(wb_dir, model_num):
	wb = load_workbook(wb_dir)
	sheet = wb.active
	currentRow = '3'
	currentCol = 'B'
	currentVal = sheet[currentCol+currentRow].value
	#lowest reference character that can be used
	intialRefChar = 'B'
	#get the ascii val of char and increase it by model number
	ref_char = chr(ord(intialRefChar)+(model_num-1))
	#typically the last 4 digits of a cell with a ref contains the cell ref
	#ex: $D$2
	ref_char_index = -3
	
	#checks if the design table name is included within sheet
	if currentVal is None:
		currentRow = '2'
		currentVal = sheet[currentCol+currentRow].value

	if type(currentVal) is not str or currentVal.find('=') == -1:
		create_log(wb_dir.split('\\')[-1] + ' does not contain a formula in the expected cell. Skipped update operation.')
		return False
	
	#check if formula includes a $ where the cell is referenced
	if sheet[currentCol+currentRow].value.split('!')[-1].find('$') == -1:
		ref_char_index = -2

	while sheet[currentCol+currentRow].value is not None:
		currentVal = sheet[currentCol+currentRow].value
		sheet[currentCol+currentRow].value = currentVal[:ref_char_index] + ref_char + currentVal[ref_char_index + 1:]
		currentCol = chr(ord(currentCol)+1)

	wb.save(wb_dir)
	wb.close()
	return

def check_reference(wb_dir):

	wb = load_workbook(wb_dir)
	sheet = wb.active
	currentRow = '3'
	currentCol = 'B'
	#lowest reference character that can be used
	intialRefChar = 'B'

	#typically the last 4 digits of a cell with a ref contains the cell ref
	#ex: $D$2
	ref_char_index = -3

	#checks if the design table name is included within sheet
	if sheet[currentCol+currentRow].value is None:
		currentRow = '2'
	
	if type(sheet[currentCol+currentRow].value) is not str or sheet[currentCol+currentRow].value.find('=') == -1:
		return 0

	#check if formula includes a $ where the cell is referenced
	if sheet[currentCol+currentRow].value.split('!')[-1].find('$') == -1:
		ref_char_index = -2
	
	ref_char = sheet[currentCol+currentRow].value[ref_char_index]
	
	#returns model number referenced. Model numbers are base off of the column value or character
	return ord(ref_char) - ord(intialRefChar) + 1

def run_macro(workbook_name, com_instance):

	wb = com_instance.workbooks.open(workbook_name)
	com_instance.AskToUpdateLinks = False

	try:
		#recalculate links
		wb.UpdateLink(Name = wb.LinkSources())

	except Exception as e:
		create_log(str(e))

		wb.Close(True)
		wb = None
		return False

	wb.Close(True)
	wb = None
	return True

def run_excel(workbook_dirs):

	xl_app = Dispatch("Excel.Application")
	xl_app.Visible = False
	xl_app.DisplayAlerts = False
	updated_wbs = 0

	for wb_dir in workbook_dirs:
		if run_macro(wb_dir, xl_app) is True:
			updated_wbs += 1
		
	xl_app.Quit()
	xl = None

	return updated_wbs

if __name__ == "__main__":

	dir_list = get_directoires()
	workbook_dirs = get_workbooks(dir_list)
	total_wbs = len(workbook_dirs)
	modified_wbs = 0
	model_num = 0

	if workbook_dirs == []:
		exit_message('No excel files found within the listed directories.\nClosing application', 4)
	else:
		create_log('Found ' + str(total_wbs) + ' workbook(s) to update')

		for wb in workbook_dirs: #check valid workbooks for model number currently being used
			model_num = check_reference(wb)
			if model_num == 0:
				continue
			break

		if model_num == 0:
			exit_message("None of the workbooks found within the given directories are in a format solidworks would recognize or they don't contain linked references.\nTry re-editing directories.txt and launch the application again", 5)

		print("Parts are currently based off of the measurement from model " + str(model_num))
		if input("Base design off other CPA model? Type y or n: ").lower() == 'y':
			while True:
				try:
					model_num = int(input('Type model number to reference: '))
					if model_num != 0:
						break
					print("Model number can't be 0")
				except ValueError:
					print("Model number cannot contain letters. Try again")
					continue

			print('Changing linked references...')

			#[:] creates copy of array
			for wb_dir in workbook_dirs[:]:
				if modifyFormula(wb_dir, model_num) is not False:
					modified_wbs += 1
				else:
					#remove from list of directories so that it doesn't get updated
					workbook_dirs.remove(wb_dir)
					
	create_log(str(modified_wbs) + ' out of ' + str(total_wbs) + ' workbooks modified')
	
	if input("Update linked references? Type y or n: ").lower() == 'y':
		create_log(str(run_excel(workbook_dirs)) + ' out of ' + str(len(workbook_dirs)) + ' workbooks updated')

	exit_message(str(time.strftime("Completed operations at %m-%d %H:%M",time.localtime())), 5)
